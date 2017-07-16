from bs4 import BeautifulSoup
import requests
import pyexcel as pe
import xlwt
import openpyxl
import json
from datetime import datetime
import time
import random
from tkinter import *

baseurl = 'https://realt.by/rent/restorant-cafe/' # Базовый URL  - https://realt.by/sale/shops/
headers = {
    'User-Agent' : 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/59.0.3071.115 Safari/537.36'
      }
excel_path = "Форма_04_Предложения_продажи_и_аренды.xlsx"
excel_sheet = "Предложения"
column_num = 70 # Количество столбов в екселе
html_folder = 'HTMLs'# Название папки, куда сохранять файлы хтмл




with open('Offices_Realt_Excel', 'r', encoding='utf-8') as jf: #открываем файл на чтение
    Realt_Excel_dict = json.load(jf) # загружаем из файла данные в словарь Realt_Excel_dict = {'Вид объекта': 'Наименование', 'Вид объекта2': 'Назначение', 'Условия сделки': 'Тип предложения', ...
excel_fields_list = list(Realt_Excel_dict.values()) # Cоздаем лист с полями Ексель - ['Наименование', 'Назначение', 'Тип предложения', 'Контактные данные'...
realt_fields_list = list(Realt_Excel_dict.keys())

with open('Offices_Realt_Fields_Options', 'r', encoding='utf-8') as jf: #открываем файл на чтение
    Excel_options_dict = json.load(jf)


def get_html(url): # Получаем html. На вход url страницы
    try:
        res = requests.get(url, headers = headers)
    except requests.ConnectionError:
        return
    if res.status_code < 400:
        return res.content


def get_dol_kurs(url = 'http://www.nbrb.by/API/ExRates/Rates/145'): # Получаем курс доллара На вход - запрос апи с сайта нацбанка
    byte_kurs = get_html(url) # b'{...}
    dict_kurs = json.loads(byte_kurs) # {...}
    kurs = dict_kurs['Cur_OfficialRate']
    return kurs # 1.9750


def get_today_date(): # Получаем егодняшнюю дату в формате '13.07.2017'
    date = datetime.strftime(datetime.now(), "%d.%m.%Y")
    return date


def del_space(string): # Для удаления спец пробела в ЦЕНЕ - 1 670. Возвращает число в формате float без пробела. На вход - строка с ценой '1 670' or '879'
    string = str(string)
    if ' ' in string:
        new_string = string.replace(' ', '')
        string = float(new_string)
        return string
    else:
        string = float(string)
        return string

def del_coma(string): # Для удаления запятой  в ЦЕНЕ - 5,76 и замене ее на точку. Возвращает строку
    if ',' in string:
        string = string.replace(',', '.')
        return string
    else:
        return string

def get_coords(soup, project): # Для каждого объявления находит координату с Яндекс карты, На вход - объект суп страницы с объявлением и словарь, куда записываются координаты
    table = soup.find_all('script', {'type': 'text/javascript'}) # Находим все объекты на странице с заданными параметрами
    for i in table:
        string = i.text
        if 'ymaps' in string: # Если в строке есть 'ymaps' - то данный скрипт содержит координаты
            coordinates = string.split('center: [')[1].split(']')[0]
            X = coordinates.split(', ')[0]
            Y = coordinates.split(', ')[1]
            project['XCoord'] = X
            project['YCoord'] = Y


def write_webpage_to_html(html_number, html_obj, folder_path):  # Сохраняет веб-страницу в файл формата .html. На вход - уникальный номер объявления (для формирования имени), html_obj = get_html(obj_url), где obj_url - url странички, которую записывают в html-файл и folder_path - путь к папке, куда сохранять файлы
    name_html = '{}/{}.html'.format(folder_path, html_number)
    with open(name_html, 'wb') as file:
        file.write(html_obj)


def get_photos(soup, project, folder_path): # НЕ ИСПОЛЬЗУЕТСЯ Загружает фотографии со страницы с объявлением. На вход soup - объект суп для странички с объявлением (В данном случае - soup1), project - словарь, где записан уникальный номер объявления (для названия фото) и folder_path - путь к папке, куда сохранять файлы
    i = 1 # Для названия фотографий
    photos = soup.find_all('div', {'class': 'photo-item'})  # получаем список с фотографиями объявления
    if photos:  # если фото на странице есть
        for photo in photos:
            lnk = photo.find('img').get('src') # Получаем ссылку на каждое фото
            nametemp = "{}/{}_{}.jpeg".format(folder_path, project['ID_object'], i)
            i +=1
            with open(nametemp, "wb") as f:
                real_photo = get_html(lnk)
                f.write(real_photo)

def get_table_coords(project, i): # НЕ ИСПОЛЬЗУЕТСЯ Для объявлений, где есть координаты в таблице на странице - для Общественно - деловой зоны их нет. На вход - project - словарь, куда записываются координты,  i - (for i in table, где table - список со всеми полями и ответами со старнички с объявлением) table = soup1.find_all('tr', {'class': 'table-row'})
    if 'Координаты для онлайн карт' in i.text:
        coordinates = i.text.split('Координаты для онлайн карт')[1].strip()
        project['Xcoord'] = coordinates.split(' ')[0]
        project['Ycoord'] = coordinates.split(' ')[1]

def write_projects_into_new_excel(projects, options, excel_path): # НЕ ИСПОЛЬЗУЕТСЯ записывает список словарей в новый файл ексель, projects - список словарей, где ключ - название поля, значение - значение для строки;  options  - список значений полей для ексель; excel_path - полный путь к новому екселю
    wb = xlwt.Workbook()
    ws = wb.add_sheet(excel_sheet)
    row = 0
    line = 0
    for option in options:
        ws.write(row, line, option)
        line += 1
    for project in projects:
        line = 0
        row += 1
        for option in options:
            if option in project:
                ws.write(row, line, project[option])
                line += 1
    wb.save(excel_path)

# 3 ФУНКЦИИ для добавления в существующий ексель

def write_into_cell(ws, project, row_num): # походится по столбцам екселя ищет совпадение с ключами в словаре project, если находит - то записывает в следующую строку значение
    for i in range(1, column_num):  # column_num - количество столбцов в екселе максимальное)))
        # for field in excel_fields_list:
        for field in list(project.keys()):  # проходимся по полям которые есть в конкретном project
            if field == ws.cell(row=3,
                                column=i).value and field == "ПДФ":  # и если поле в project совпадает с полем в ексель и это поле для гиперссылки на файл html,  то записываем в строку гиперссылку
                ws.cell(row=row_num, column=i).hyperlink = project[field]
            elif field == ws.cell(row=3,
                                  column=i).value:  # и если поле в project совпадает с полем в ексель то записываем в строку значение по ключу в project
                ws.cell(row=row_num, column=i).value = project[field]

def add_projects_into_existing_excel(projects, excel_path = "MyExcel.xlsx"): # записывает все projects (лист со влоарями) в ексель, т.е. все объявления со страницы
    wb = openpyxl.load_workbook(filename=excel_path)# открываем существующий ексель
    ws = wb.get_sheet_by_name(excel_sheet) # Выбираем лист
    row_num = ws.max_row # находим последнюю строку (чтобы записывать новые данные в следующую)
    for project in projects:
        row_num +=1
        write_into_cell(ws, project, row_num)
    wb.save(excel_path)

def add_project_into_existing_excel(project, excel_path = "MyExcel.xlsx"): # записывает один project (словарь с одним объявлением) в ексель
    wb = openpyxl.load_workbook(filename=excel_path) # открываем существующий ексель
    ws = wb.get_sheet_by_name(excel_sheet) # Выбираем лист
    row_num = ws.max_row  # находим последнюю строку (чтобы записывать новые данные в следующую)
    row_num += 1
    write_into_cell(ws, project, row_num)
    wb.save(excel_path)



def get_area(realt_answer): # Получаем из ответа только площадь - избавляемся от м², где realt_answer - ответ на реалте
    return realt_answer.split('м²')[0].strip()

def get_hight(realt_answer): # Получаем из ответа только площадь - избавляемся от м, где realt_answer - ответ на реалте
    realt_answer = realt_answer.split('м')[0].strip()
    return round(float(realt_answer), 2)

# 5 ФУНКЦИЙ ДЛЯ ПОЛУЧЕНИЯ ЦЕНЫ

def get_price(realt_answer, project, Excel_field): # (Если цена указана за 1 кв.м.) Функция удаляет запятую и пробел в цене если они есть и записывает в project цену переведенную в долларах с округлением до 2 знаков после запятой
    realt_answer = del_coma(realt_answer)
    realt_answer = del_space(realt_answer)
    project[Excel_field] = round((realt_answer / get_dol_kurs()), 2)

def get_price_whole_lot(realt_answer, project, Excel_field): # (Если цена указана за ВЕСЬ ЗЕМЕЛЬНЫЙ УЧАСТОК) Функция удаляет запятую и пробел в цене если они есть и записывает в project цену переведенную в долларах С ПОМЕТКОЙ ЧТО ЦЕНА УКАЗАНА ЗА ВЕСЬ ЗЕМЕЛЬНЫЙ УЧАСТОК с округлением до 2 знаков после запятой
    realt_answer = del_coma(realt_answer)
    realt_answer = del_space(realt_answer)
    project[Excel_field] = '{} !Цена указана за весь участок'.format(round((realt_answer / get_dol_kurs()), 2))

def check_price(realt_answer, project, Excel_field): # (Если цена указана за 1 кв.м.)  Функция проверяет есть ли в цене строки '—', 'до ' или нет. и записывает корректную (если были строки) в долларах
    if '—' in realt_answer: # Цена записана в таком виде: 114 911 — 120 658 руб, 1 973—2 072 руб/кв.м
        realt_answer = realt_answer.split('—')[0].strip()
        get_price(realt_answer, project, Excel_field)
    elif 'до ' in realt_answer: # Цена записана в таком виде: до 92 750 руб, до 2 157 руб/кв.м
        realt_answer = realt_answer.split('до ')[1].strip()
        get_price(realt_answer, project, Excel_field)
    else:
        get_price(realt_answer, project, Excel_field)

def check_price_whole_lot(realt_answer, project, Excel_field): # (Если цена указана за ВЕСЬ ЗЕМЕЛЬНЫЙ УЧАСТОК) Функция проверяет есть ли в цене строки '—', 'до ' или нет. и записывает корректную (если были строки) в долларах
    if '—' in realt_answer:
        realt_answer = realt_answer.split('—')[0].strip()
        get_price_whole_lot(realt_answer, project, Excel_field)
    elif 'до ' in realt_answer:
        realt_answer = realt_answer.split('до ')[1].strip()
        get_price_whole_lot(realt_answer, project, Excel_field)
    else:
        get_price_whole_lot(realt_answer, project, Excel_field)

def get_finish_price(realt_answer, project, Excel_field): # где realt_answer - ответ на реалте, project - словарь, куда записывается итоговая цена,   Excel_field = Realt_Excel_dict[option] - Получаем название поля в Excel, option - название поля на реалте
    # print(realt_answer) - сто пятьсот вариантов
    if ", " in realt_answer: # Цена записана в таком виде: 11 740 руб, 118 руб/кв.м
        realt_answer = realt_answer.split(', ')[1].split('руб')[0].strip()  # '355—395' or '356' or '1 876' or '1 355—1 395' or 'до 2 157'
        check_price(realt_answer, project, Excel_field)
    elif 'договор' in realt_answer: # Цена записана в таком виде: Цена договорная
        project[Excel_field] = 'Цена договорная'
    else:
        if ' руб/кв.м' in realt_answer: # есть цена за 1 кв.м
            realt_answer = realt_answer.split(' руб/кв.м')[0].strip()
            check_price(realt_answer, project, Excel_field)
        elif ' руб/м²' in realt_answer:
            realt_answer = realt_answer.split(' руб/м²')[0].strip()
            check_price(realt_answer, project, Excel_field)
        else: # значит стоимость дана только за весь земельный участок и в строке есть слово "руб"
            realt_answer = realt_answer.split(' руб')[0].strip()
            check_price_whole_lot(realt_answer, project, Excel_field)

# 2 ФУНКЦИИ ДЛЯ ОПРЕДЕЛЕНИЯ ВИДА, НАЗНАЧЕНИЯ И НАИМЕНОВАНИЯ ОБЪЕКТА
def write_into_project_all_vidy(osnov_vid, project, Excel_field, Excel_field2, Excel_field3):
    project[Excel_field] = Excel_options_dict[Excel_field][osnov_vid]
    project[Excel_field2] = Excel_options_dict[Excel_field2][osnov_vid]
    project[Excel_field3] = Excel_options_dict[Excel_field3][osnov_vid]

def get_finish_vid_object(realt_answer, project, Excel_field, Excel_field2, Excel_field3): # из многообразия того что записано в поле Вид объекта, нужно определить основной вид и в определить по нему Вид объекта, Наименование и Назначение
    if "(" in realt_answer: # такой вид: "Продажа офисов от застройщика в новостройке по пр.Дзержинского (офис)"
        osnov_vid = realt_answer.split(" (")[0].strip().lower() # "Продажа офисов от застройщика в новостройке по пр.Дзержинского"
        if len(osnov_vid) <= 18: # "торговое помещение" - 18 символов. Самое длинное из возможных вариантов на реалте, которое можеть быть записано до скобки. Но бывает что до скобок записана ерунда и она тоже меньше 18 символов
            try:
                write_into_project_all_vidy(osnov_vid, project, Excel_field, Excel_field2, Excel_field3)
            except KeyError:
                osnov_vid = realt_answer.split(")")[-2].split("(")[1].lower()# Нужно взять только то, что в последних скобках - Делим по последней скобке и берем предпоследний эл-т - [-2] - второй элемент с конца, т.к. первый элемент с конца - пустая строка
                if ',' in osnov_vid: # если в скобочках записано более чем один доп вид - т.е. есть запятая. ПОЧТИ ВСЕГДА
                    osnov_vid = osnov_vid.split(",")[0]
                    write_into_project_all_vidy(osnov_vid, project, Excel_field, Excel_field2, Excel_field3)
                else: # если в скобочках записан один вид
                    write_into_project_all_vidy(osnov_vid, project, Excel_field, Excel_field2, Excel_field3)
        else: # если символов до "(" большее 18, то основной вид записан первым после послдней скобки "("
            osnov_vid = realt_answer.split(")")[-2].split("(")[1].lower() # Нужно взять только то, что в последних скобках - Делим по последней скобке и берем предпоследний эл-т - [-2] - второй элемент с конца, т.к. первый элемент с конца - пустая строка
            if ',' in osnov_vid: # если в скобочках записано более чем один доп вид - т.е. есть запятая. ПОЧТИ ВСЕГДА
                osnov_vid = osnov_vid.split(",")[0]  # если в скобочках записано более чем один доп вид. ПОЧТИ ВСЕГДА
                write_into_project_all_vidy(osnov_vid, project, Excel_field, Excel_field2, Excel_field3)
            else: # если в скобочках записан один вид
                write_into_project_all_vidy(osnov_vid, project, Excel_field, Excel_field2, Excel_field3)
    else:
        osnov_vid = realt_answer.lower() # Если в поле нет скобочек, значит записан только один основной вид
        write_into_project_all_vidy(osnov_vid, project, Excel_field, Excel_field2, Excel_field3)


def get_contacts(realt_answer, project, Excel_field): # Если в поле ответа "Контактные данные" есть 'Пожалуйста, скажите что Вы нашли это объявление на сайте Realt.by', то удаляется эта часть
    if 'Пожалуйста, скажите что Вы нашли это объявление на сайте Realt.by' in realt_answer:
        project[Excel_field] = realt_answer.split('Пожалуйста, скажите что Вы нашли это объявление на сайте Realt.by')[1].strip()
    else:
        project[Excel_field] = realt_answer

# 5 функций для определения адреса
def get_street(realt_answer, project, Excel_field, Excel_field2): # Получаем название улицы
    street_elem = realt_answer.split(".")[0]  # Никольская ул
    realt_elem_name = street_elem.split(' ')[-1]  # ул
    realt_street_name = street_elem.split(realt_elem_name)[0].strip()  # 40 лет Победы
    project[Excel_field2] = Excel_options_dict[Excel_field2][realt_street_name]

def get_street_not_in_dict(realt_answer, project, Excel_field, Excel_field2): # Получаем название улицы, если ее нет в словаре Offices_Realt_Fields_Options
    street_elem = realt_answer.split(".")[0]  # Никольская ул
    realt_elem_name = street_elem.split(' ')[-1]  # ул
    realt_street_name = street_elem.split(realt_elem_name)[0].strip()  # 40 лет Победы
    project[Excel_field2] = "{} - не из классификатора".format(realt_street_name)

def get_elem(realt_answer, project, Excel_field, Excel_field2): # Получаем название ЭУДС
    street_elem = realt_answer.split(".")[0]  # Никольская ул
    realt_elem_name = street_elem.split(' ')[-1]  # ул
    project[Excel_field] = Excel_options_dict[Excel_field][realt_elem_name]

def get_elem_not_in_dict(realt_answer, project, Excel_field, Excel_field2): # Получаем название УДС, если ее нет в словаре Offices_Realt_Fields_Options
    street_elem = realt_answer.split(".")[0]  # Никольская ул
    realt_elem_name = street_elem.split(' ')[-1]  # ул
    project[Excel_field] = "{} - не из классификатора".format(realt_elem_name)

def get_full_address(realt_answer, project, Excel_field, Excel_field2, Excel_field3, Excel_field4, id_object_name): # Никольская ул., 66-2, 40 лет Победы ул., 66-2,
    if "." in realt_answer: # Сначала ищем в ответе точку - она всегда после ЭУДС (ВОПРОС - точно ли ВСЕГДА - например есть улица "С. Ковалевской" или "Меньковский тракт"
        try:
            get_street(realt_answer, project, Excel_field, Excel_field2)
        except IndexError:
            print("Для объекта с номером {} - Невозможно определить улицу".format(id_object_name))
        except KeyError: # улицы нет в словаре Offices_Realt_Fields_Options
            get_street_not_in_dict(realt_answer, project, Excel_field, Excel_field2)
        try:
            get_elem(realt_answer, project, Excel_field, Excel_field2)
        except IndexError:
            print("Для объекта с номером {} - Невозможно определить улицу".format(id_object_name))
        except KeyError:
            get_elem_not_in_dict(realt_answer, project, Excel_field, Excel_field2)
    else: # обработка если в ЭУДС нет точки - например Меньковский тракт
        if "," in realt_answer: # Меньковский тракт, 43
            try:
                street_elem = realt_answer.split(",")[0]
                realt_elem_name = street_elem.split(' ')[-1]  # ул
                realt_street_name = street_elem.split(realt_elem_name)[0].strip()  # 40 лет Победы
                project[Excel_field2] = Excel_options_dict[Excel_field2][realt_street_name]
                project[Excel_field] = Excel_options_dict[Excel_field][realt_elem_name]
            except KeyError: # сделано для адреса Центральная, 13
                print("Неправильная структура УДС") # сделано для адреса "Центральная, 13
        else: # Меньковский тракт
            try:
                realt_elem_name = realt_answer.split(' ')[-1]  # ул
                realt_street_name = realt_answer.split(realt_elem_name)[0].strip()  # 40 лет Победы
                project[Excel_field2] = Excel_options_dict[Excel_field2][realt_street_name]
                project[Excel_field] = Excel_options_dict[Excel_field][realt_elem_name]
            except KeyError: # сделано для адреса "Брест"
                print("Неправильная структура УДС")
    if "," in realt_answer: # если в ответе есть запятая, значит указаны номер дома/корпус
        house_korp = realt_answer.split(',')[1].strip()  # 66-2 или 66 or "66-2 Информация о доме"
        if "-" in house_korp:  # значит есть корпус
            house = house_korp.split('-')[0]
            project[Excel_field3] = int(house)
            korp = house_korp.split('-')[1].strip()
            if " " in korp: # если после адреса есть строка "Информация о доме"
                korp = korp.split(' ')[0]
            project[Excel_field4] = korp # корпус может быть не только integer
        else:# значит нет корпуса
            house = house_korp
            if " " in house_korp:# если после адреса есть строка "Информация о доме"
                house = house.split(' ')[0]
            project[Excel_field3] = int(house)

def parse_object(obj_url, project={}): # Парсим одно объявление
    html_obj = get_html(obj_url)
    soup = BeautifulSoup(html_obj, "html.parser")
    table = soup.find_all('tr', {'class': 'table-row'})  # Получаем список со всеми необходимыми данными объявления

    id_object_name = int(
        obj_url.split('object/')[1][:-1])  # из url страницы объявления оставляем только уникальный номер. Из этого https://realt.by/sale/offices/object/712024/ - получаем 712024
    project['№ Объявления'] = id_object_name
    project['Дата актуальности предложения'] = get_today_date()  # Сегодняшняя дата (13.07.2017)
    project['ПДФ'] = '{}/{}.html'.format(html_folder, id_object_name)
    project['Источник'] = "Realt.by"
    get_coords(soup, project)  # Координаты Х и У записываются в словарь project

    # Для бизнес-центров и торговых центров в объявлении нет строки "Вид объекта". ПОэтому записываем его таким образом
    if 'malls' in baseurl: # Если парсим Торговые центры
        Excel_field = Realt_Excel_dict['Вид объекта']
        Excel_field2 = Realt_Excel_dict['Вид объекта2']
        Excel_field3 = Realt_Excel_dict['Вид объекта3']
        osnov_vid = "торговый центр"
        write_into_project_all_vidy(osnov_vid, project, Excel_field, Excel_field2, Excel_field3)

    if 'newoffices' in baseurl: # Если парсим Бизнес-центры
        Excel_field = Realt_Excel_dict['Вид объекта']
        Excel_field2 = Realt_Excel_dict['Вид объекта2']
        Excel_field3 = Realt_Excel_dict['Вид объекта3']
        osnov_vid = "бизнес-центр"
        write_into_project_all_vidy(osnov_vid, project, Excel_field, Excel_field2, Excel_field3)

    write_webpage_to_html(id_object_name, html_obj,
                          html_folder)  # write web page to html file

    for i in table:  # Проходимся по каждой строке на странице объявления

        option = i.find('td', {'class': "table-row-left"}).text # в строке выделяем левую часть - т.е. название поля
        if option in realt_fields_list: # если полю на реалте есть соответствие в словаре Offices_Realt_Excel - значит его обрабатываем, т.к. в объявлении есть лишние поля, которые нам не нужны
            realt_answer = i.find('td', {'class': "table-row-right"}).text # в строке выделяем правую часть - т.е. ответ
            print(option)
            print(realt_answer)
            Excel_field = Realt_Excel_dict[option]  # Получаем название поля в Ексель, соответствующее полю на реалте

            if option == "Площадь":
                project[Excel_field] = get_area(realt_answer)

            elif option == "Ориентировочная стоимость эквивалентна":
                # realt_answer = 1 677 руб/кв.м 1 677 руб/кв.м  Цена сделки определяется по соглашению сторон. Расчеты осуществляются в белорусских рублях в соответствии с законодательством Республики Беларусь.
                get_finish_price(realt_answer, project, Excel_field)

            elif option == "Вид объекта":
                Excel_field2 = Realt_Excel_dict['Вид объекта2']
                Excel_field3 = Realt_Excel_dict['Вид объекта3']
                get_finish_vid_object(realt_answer, project, Excel_field, Excel_field2, Excel_field3)

            # Старый код, Удалить если всё будет ОК. чтобы работал, в словаре Offices_Realt_Fields_Options в вариантах ответа реалта нужно удалить НДС

            # elif option == "НДС":
            #     try:
            #         realt_answer = i.text.split(option)[
            #             2].strip()  # потому что realt_answer гзначально такой список ['', ' ', ' не включен)'] Поэтому берем третий элемент (Также поменяла в json словаре Offices_Realt_Fields_Options.json
            #         project[Excel_field] = Excel_options_dict[Excel_field][realt_answer]
            #     except:  # Может быть что поля НДС нет на странице но строка "НДС встречается в тексте. Поэтому если она встречается то просто пропустить
            #         pass

            elif option == "Телефоны":
                get_contacts(realt_answer, project, Excel_field)

            elif option == "Вода":
                Excel_field2 = Realt_Excel_dict['Вода2'] # Горяее водоснабжение
                project[Excel_field] = Excel_options_dict[Excel_field][realt_answer]
                project[Excel_field2] = Excel_options_dict[Excel_field2][realt_answer]

            elif option == "Высота потолков":
                project[Excel_field] = get_hight(realt_answer)

            elif option == "Адрес":  # Никольская ул., 66-2, 40 лет Победы ул., 66-2,
                Excel_field2 = Realt_Excel_dict['Адрес2']  # название улицы
                Excel_field3 = Realt_Excel_dict['Адрес3']  # номер дома
                Excel_field4 = Realt_Excel_dict['Адрес4']  # корпус
                get_full_address(realt_answer, project, Excel_field, Excel_field2, Excel_field3, Excel_field4,
                                 id_object_name)
            elif option == "Район области":
                project[Excel_field] = realt_answer.split('район')[0].strip()

            elif option == "Этажность":
                realt_answer = realt_answer.split(' этажей')[0]
                if '-' in realt_answer:
                    realt_answer = realt_answer.split('-')[1].strip()
                project[Excel_field] = realt_answer

            elif Realt_Excel_dict[option] in Excel_options_dict:  # доп действий производить не нужно, ответ записан в той форме, в которой он в словаре Offices_Realt_Fields_Options
                try:
                    project[Excel_field] = Excel_options_dict[Excel_field][realt_answer]
                except KeyError:  # есть объявление https://realt.by/sale/shops/object/1106690/ у которого Материал стен не из классификатора
                    project[Excel_field] = '{} - не из классификатора'.format(realt_answer)
            else: # Записываем ответ как он есть
                project[Excel_field] = realt_answer
    print(project)
    return project


def parse_page(html): # Парсим одну страницу
    soup = BeautifulSoup(html, "html.parser")
    # look for hrefs in titles
    table = soup.find_all('div', {'class': 'bd-item'}) # Получаем список с объявлениями на странице

    projects = [] # В дальнейшем - Список со словарями со страницы, где каждый словарь соответствует одному объявлению на странице. Словарь в таком виде, в котором будет записываться в Ексель
    # проходимся по каждому объявлению
    for row in table:
        # получаем ссылку каждого объявления
        href_name = row.find('a')
        obj_url = href_name.get("href")
        project = {} # Для записи одного объявления в словарь
        parse_object(obj_url, project) # Начинаем парсить каждое оюъявление. Результат - project
        projects.append(project)# Добавляем словарь с днными по каждому оюъявлению в лист
    print(projects)
    return projects

# ДЛЯ ПРОСМОТРА ОДНОГО ОБЪЯВЛЕНИЯ
# baseurl = 'https://realt.by/sale/offices/object/1059798/'
# parse_object(baseurl)

# Проба для записи только одного объявления в ексель
# baseurl = 'https://realt.by/sale/offices/object/1059798/'
# add_project_into_existing_excel(parse_object(baseurl), excel_path = excel_path)

# ПОШЛА РАБОТА
# Получаем хтмл контент базового урла
html = get_html(baseurl)
# Записываем главную страницу (первую базового урла) в ексель
add_projects_into_existing_excel(parse_page(html), excel_path=excel_path)

# Дальше работа со следующими страницами для базового урла. Находим количество страниц и двигаемся по ним
soup = BeautifulSoup(html, "html.parser")
pages = soup.find('div', {'class': 'uni-paging'})  # Находим тег со страницами
if pages:
    last_page = pages.text.split("... ")[1].strip() # 1, 2, 3, 4... 78
    print("The number of pages: {}".format(last_page))
    last_url = int(last_page) - 1 # для второй страницы last_url = 1, поэтому для последней last_page - 1
    # ВЫБИРАЕМ С КАКОЙ ПО КАКУЮ СТРАНИЦУ ПАРСИТЬ, range(1, 3) - [1, 2] - 1 - вторая страница, 3 - третья, т.к. 3 - не входит
    for i in range(1, 3):
        url = "{}?page={}".format(baseurl, i)
        print(url)
        html = get_html(url)
        add_projects_into_existing_excel(parse_page(html), excel_path=excel_path)
        the_last_succesful_page = 1
        print("The last succesful page is {}".format(i+1))
        waiting_time = random.randint(1, 10)
        print("Waiting time is {}".format(waiting_time))
        time.sleep(waiting_time)







