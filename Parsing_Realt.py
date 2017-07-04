from bs4 import BeautifulSoup
import requests
import pyexcel as pe
import xlwt
import openpyxl
import os

baseurl = 'https://realt.by/sale/shops/?page=1'
field_dict = {'Канализация': 'Sewerage', 'Электричество': 'Electricity', 'Дата обновления': 'ObjectData', 'Телефоны': 'Contacts',
              'E-mail': 'E-mail', 'Контактное лицо': 'ContactName', 'Область': 'Область', 'Район области': 'Район области', 'Населенный пункт': 'Населенный пункт', 'Направление': 'Direction', 'Адрес': 'Адрес',
             'Вид объекта': 'ObjectType','Площадь участка': 'ZUArea', 'Площадь': 'Area', 'Этаж / этажность': 'Floor', 'Высота потолков': 'Hihgt', 'Материал cтен': 'WallMaterial',
             'Год постройки': 'BuildYear', 'Состояние здания / помещения': 'Состояние', 'Кол-во помещений': 'RoomNumber', 'Кол-во телефонов': 'PhoneAmount', 'Наличие оборудования': 'Наличие оборудования',
             'Естественное освещение': 'NaturalLight', 'Отопление': 'Heat', 'Электроснабжение': 'Электроснабжение', 'Подведенная мощность': 'Power', 'Газоснабжение': 'Gas', 'Вода': 'Water',
             'Сан.узел': 'Toilet', 'Юридический адрес': 'LegalAddress',
             'Телефон': 'Phone','Дополнительно': 'Additionally', 'Примечания': 'Примечания','Условия продажи': 'SaleConditions', 'Вид владения': 'Вид владения', 'Ориентировочная стоимость эквивалентна': 'Price'}
options = list(field_dict.keys())
fields = list(field_dict.values())
my_fields = ['ID_object', 'Xcoord', 'Ycoord']
fields.extend(my_fields)
print(fields)

def get_html(url):
    try:
        res = requests.get(url)
    except requests.ConnectionError:
        return

    if res.status_code < 400:
        return res.content

def parse(html):
    soup = BeautifulSoup(html, "html.parser")
    # look for hrefs in titles
    table = soup.find_all('div', {'class': 'bd-item'})
    projects = []
    # проходимся по каждому объявлению
    for row in table:
        i=1 # for name of photo
        # get hrefs of all pages
        href_name = row.find('a')
        obj_url = href_name.get("href")
        html_obj = get_html(obj_url)

        soup1 = BeautifulSoup(html_obj, "html.parser")
        table = soup1.find_all('tr', {'class': 'table-row'})
        project = {}
        project['ID_object'] = int(obj_url.split('object/')[1][:-1])

        # # download photos
        # photos = soup1.find_all('div', {'class': 'photo-item'})
        # print(photos)
        # if photos:
        #     for photo in photos:
        #         print(photo)
        #         lnk = photo.find('img').get('src')
        #         print(lnk)
        #         nametemp = "{}_{}.jpeg".format(project['ID_object'], i)
        #         print(nametemp)
        #         i+=1
        #         with open(nametemp, "wb") as f:
        #             f.write(requests.get(lnk).content)


        for i in table:
            if 'Координаты для онлайн карт' in i.text:
                coordinates = i.text.split('Координаты для онлайн карт')[1].strip()
                project['Xcoord'] = coordinates.split(' ')[0]
                project['Ycoord'] = coordinates.split(' ')[1]
            for option in options:
                if option in i.text:
                    project[field_dict[option]] = i.text.split(option)[1].strip()
        # print(project)
        projects.append(project)
    print(projects)

    # Write into NEW EXCEL. NOW NOT USED

    # wb = xlwt.Workbook()
    # ws = wb.add_sheet("Sheet")
    # row = 0
    # line = 0
    # for option in options:
    #     ws.write(row, line, option)
    #     line += 1
    # for project in projects:
    #     line = 0
    #     row += 1
    #     for option in options:
    #         if option in project:
    #             ws.write(row, line, project[option])
    #             line += 1
    # wb.save("D:/TESTYYYY.xls")

    # !!!! NOW USE - WRITE TO EXISTING EXCEL

    file = "D:/TESTYYYY1.xlsx"
    wb = openpyxl.load_workbook(filename=file)
    # Seleciono la Hoja
    ws = wb.get_sheet_by_name('Sheet')
    row_num = ws.max_row
    print(row_num)
    for project in projects:
        row_num += 1
        for i in range(1, 70):
            for field in fields:
                if field == ws.cell(row=1, column=i).value and field in project:
                    ws.cell(row=row_num, column=i).value = project[field]
    wb.save(file)


html = get_html(baseurl)
parse(html)


# soup = BeautifulSoup(html, "html.parser")
# pages = soup.find('div', {'class': 'uni-paging'})
# last_page = pages.text.split("... ")[1].strip()
# print(last_page)
#
# last_url = int(last_page) - 1
# for i in range(1, last_url):
#     url = "{}?page={}".format(baseurl, i)
#     html = get_html(url)
#     parse(html)